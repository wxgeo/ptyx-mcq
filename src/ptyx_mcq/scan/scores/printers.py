import csv
import math
from abc import abstractmethod, ABC
from collections import Counter
from itertools import chain
from typing import TYPE_CHECKING, Callable, Iterable

from openpyxl import Workbook
from openpyxl.cell import Cell, MergedCell
from openpyxl.chart import Reference, BarChart
from openpyxl.chart.layout import ManualLayout, Layout
from openpyxl.comments import Comment
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.worksheet import Worksheet

from ptyx.pretty_print import TermColors, bold, green, red, term_color, yellow, print_info
from ptyx_mcq.tools.parse_config.subtypes import OriginalQuestionNumber

if TYPE_CHECKING:
    from ptyx_mcq.scan.scores.scores_manager import ScoresManager

STATISTIC_INFO = {"mean": "AVERAGE", "min": "MIN", "max": "MAX"}


class ScoresPrinter(ABC):
    def __init__(self, parent: "ScoresManager"):
        super().__init__()
        self.parent = parent

    @staticmethod
    def _convert(num: float | str, factor=1.0):
        if isinstance(num, float):
            num = round(factor * num, 2)
        return num

    @abstractmethod
    def run(self): ...


class TerminalScoresPrinter(ScoresPrinter):
    def run(self) -> None:
        min_score: float = math.inf
        max_score: float = -math.inf
        print()
        print(term_color(f"SCORES (/{self.parent.max_score:g}):", color=TermColors.CYAN))
        for (student_name, student_id), score in sorted(self.parent.class_scores.items()):
            score = self._convert(score)
            if isinstance(score, (float, int)):
                if score < min_score:
                    min_score = score
                if score > max_score:
                    max_score = score
            print(f" - {student_name}: {self._convert(score)}")
        if self.parent.work_scores:
            mean = round(sum(self.parent.work_scores.values()) / len(self.parent.work_scores), 2)
            print(yellow("Mean: " + bold(f"{mean:g}") + f"/{self.parent.max_score:g}"))
            print(f"Min: {red(min_score)} - Max: {green(max_score)}")
        else:
            print("No score found !")
        print()


class SheetsScoresPrinter(ScoresPrinter, ABC):
    def _write_scores(self, writerow: Callable) -> None:
        # max_score is the maximal theoretical score, not the highest actually obtained.
        max_score = self.parent.max_score
        writerow(("ID", "Name", f"Score/{max_score:g}", "Score/20", "Score/100"))
        for (student_name, student_id), score in sorted(self.parent.class_scores.items()):
            # TODO: Add ability to change the notation system.
            writerow(
                [
                    student_id,
                    student_name,
                    self._convert(score),
                    self._convert(score, factor=20 / max_score) if max_score else 0,
                    self._convert(score, factor=100 / max_score) if max_score else 0,
                ]
            )


class CsvScoresPrinter(SheetsScoresPrinter):
    def run(self) -> None:
        scores_path = self.parent.mcq_parser.scan_data.files.csv_scores
        with open(scores_path, "w", newline="") as csvfile:
            self._write_scores(writerow=csv.writer(csvfile).writerow)
        print_info(f'Results stored in "{scores_path}"')


class ExcelScoresPrinter(SheetsScoresPrinter):
    def __init__(self, parent: "ScoresManager"):
        super().__init__(parent)
        self.workbook = Workbook()

    def run(self) -> None:
        wb = self.workbook
        # grab the active worksheet
        ws: Worksheet | None = wb.active
        assert ws is not None
        self._add_main_table(ws)
        # Add a chart to illustrate scores' distribution.
        ws = wb.create_sheet(title="Scores Distribution")
        assert ws is not None
        self._add_chart(ws)
        # Add the detailed scores, question per question.
        ws = wb.create_sheet(title="Details")
        assert ws is not None
        self._add_details(ws)
        wb.save(self.parent.mcq_parser.scan_data.files.xlsx_scores)

    def _create_table(
        self,
        sheet: Worksheet,
        name: str,
        min_row=1,
        min_col=1,
        max_row: int | None = None,
        max_col: int | None = None,
    ) -> None:
        if max_row is None:
            max_row = sheet.max_row
        if max_col is None:
            max_col = sheet.max_column

        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)

        # Row Stripes: Light Blue background
        row_stripe_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        # ---------------------------------------

        # 2. Iterate and apply Direct Formatting (for LibreOffice)
        # We iterate from row 1 (header) to max_row
        for row in sheet.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
            for cell in row:
                assert cell.row is not None
                # Apply Header Style
                if cell.row == min_row:
                    cell.fill = header_fill
                    cell.font = header_font
                # Apply Striped Row Style (Even rows)
                elif (cell.row - min_row) % 2 == 1:
                    cell.fill = row_stripe_fill

        tab = Table(
            displayName=name,
            ref=f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}",
        )

        # Add a default style with striped rows and banded columns
        # noinspection PyTypeChecker
        tab.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=True,
        )
        sheet.add_table(tab)

    def _add_main_table(self, sheet: Worksheet) -> None:
        """Add the main table with the global students scores."""
        sheet.title = "Resume"
        self._write_scores(writerow=sheet.append)

        self._create_table(sheet, "StudentsScores")

        # Fix columns' widths.
        self._adjust_col_width(sheet, "A", [student_id for _, student_id in self.parent.class_scores])
        self._adjust_col_width(sheet, "B", [student_name for student_name, _ in self.parent.class_scores])

        # Append some statistics concerning the students' scores.
        n = len(self.parent.class_scores)
        self._append_stats(sheet, n + 3, 2, (3, 4, 5), (2, n + 1))

    @staticmethod
    def _adjust_col_width(sheet: Worksheet, col_ref: str, items: list) -> None:
        sheet.column_dimensions[col_ref].width = 1.23 * max([len(str(item)) for item in items], default=0)

    @staticmethod
    def _append_stats(
        sheet: Worksheet,
        first_row: int,
        first_col: int,
        cols: Iterable[int],
        row_range: tuple[int, int],
    ):
        """
        Add the average, the min and the max values at the bottom of the table.

        (`first_row`, `first_cell`) is the position of the top left cell of the stats section.
        `cols` and `row_range` refer to the data to sum-up.
        """
        for row, (legend, formula) in enumerate(STATISTIC_INFO.items(), start=first_row):
            cell = sheet.cell(row, first_col)
            cell.value = legend.capitalize()
            cell.font = Font(bold=True)
            for col in cols:
                col_letter = get_column_letter(col)
                cell = sheet.cell(row, col)
                cell.value = f"={formula}({col_letter}{row_range[0]}:{col_letter}{row_range[1]})"
                cell.number_format = "0.00"
                cell.font = Font(bold=True)

    def _add_details(self, sheet: Worksheet) -> None:
        """Add a table with the score for each question detailed."""
        scan_data = self.parent.mcq_parser.scan_data
        cfg = self.parent.mcq_parser.config
        sheet.cell(1, 1, "Student")
        # Enumerate all the questions (each version of version is seen as a different question).
        questions_seen: list[OriginalQuestionNumber] = sorted(
            set(q for doc in scan_data for q in doc.questions)
        )
        questions_col: dict[OriginalQuestionNumber, int] = {q: i + 2 for i, q in enumerate(questions_seen)}

        def _tag_cell(_cell: Cell | MergedCell, q: OriginalQuestionNumber) -> None:
            if q_name := cfg.questions_names.get(q, ""):
                _cell.comment = Comment(text=q_name, author=f"Q{q}")

        # Generate the table of the scores.
        row = 1
        for row, doc in enumerate(scan_data.sorted_by("student_name"), start=2):
            sheet.cell(row, 1, doc.student_name)
            for q, question in doc.questions.items():
                col = questions_col[q]
                header_cell = sheet.cell(1, col)
                header_cell.value = f"Q{q}"
                _tag_cell(header_cell, q)
                cell = sheet.cell(row, col)
                cell.value = question.score
                cell.number_format = "0.00"
        self._create_table(sheet, "ScoresDetails")

        # Add the weight of each question, and some statistics.
        cell = sheet.cell(row + 2, 1)
        cell.value = "Weight"
        cell.font = Font(italic=True)
        self._append_stats(sheet, row + 3, 1, range(2, len(questions_seen) + 2), (2, row))
        for q in questions_seen:
            col = questions_col[q]
            cell = sheet.cell(row + 2, col)
            weight = cfg.weight.get(q, cfg.weight["default"])
            correct = cfg.correct.get(q, cfg.correct["default"])
            skipped = cfg.skipped.get(q, cfg.skipped["default"])
            incorrect = cfg.incorrect.get(q, cfg.incorrect["default"])
            cell.value = weight
            cell.font = Font(italic=True)
            if weight > 0:
                cell_ref = f"{get_column_letter(col)}{row + 3}"
                self._color_cell_according_to_value(
                    sheet, cell_ref, min_=weight * incorrect, med_=weight * skipped, max_=weight * correct
                )
                _tag_cell(sheet[cell_ref], q)
            else:
                for i in chain(range(2, row + 1), range(row + 2, row + 6)):
                    sheet.cell(i, col).fill = PatternFill(fill_type="solid", fgColor="DDDDDD")
        # Adjust the width of the first column (students names).
        self._adjust_col_width(sheet, "A", [student_name for student_name, _ in self.parent.class_scores])

    @staticmethod
    def _color_cell_according_to_value(
        sheet: Worksheet, cell_ref: str, min_: float, med_: float, max_: float
    ):
        """
        Apply a color to the given cell, depending on its current value.

        The cell will be red if the value is `min_`, green if the value is `max_`,
        else a yellowish intermediate color, proportionally to its value.
        """
        rule = ColorScaleRule(
            start_type="num",
            start_value=min_,
            start_color="ebbbb5",  # red
            mid_type="num",
            mid_value=med_,
            mid_color="ebe5b5",  # yellow
            end_type="num",
            end_value=max_,
            end_color="bfebb5",  # green
        )
        sheet.conditional_formatting.add(cell_ref, rule)

    def _add_chart(self, sheet: Worksheet) -> None:
        """Append a chart representing the scores' distribution as a new sheet."""

        score_counts = Counter(
            round(score) for score in self.parent.class_scores.values() if isinstance(score, (int, float))
        )

        sheet.append(["Score", "Count"])
        for score in range(math.ceil(self.parent.max_score) + 1):
            sheet.append([score, score_counts.get(score, 0)])

        n = sheet.max_row
        counts_ref = Reference(sheet, min_col=2, min_row=2, max_row=n)
        scores_ref = Reference(sheet, min_col=1, min_row=2, max_row=n)

        # ✅ Use BarChart instead of LineChart
        chart = BarChart()
        chart.type = "col"  # vertical columns (not horizontal bars)
        chart.grouping = "clustered"
        chart.title = "Scores Distribution"
        chart.x_axis.title = "Score"
        chart.y_axis.title = "Number of Students"

        chart.add_data(counts_ref, titles_from_data=False)
        chart.set_categories(scores_ref)

        chart.legend = None
        chart.width = 21
        chart.height = 9
        chart.layout = Layout(
            manualLayout=ManualLayout(x=0.005, y=0.05, w=0.75, h=0.8, xMode="factor", yMode="factor")
        )

        # ✅ Set bar fill to solid red (replaces LineProperties)
        chart.series[0].graphicalProperties.solidFill = "FA8C84"
        chart.series[0].graphicalProperties.line.solidFill = "FF0000"  # bar border color

        # ✅ Control spacing between bars (lower % = narrower gaps, default is 150)
        chart.gapWidth = 50

        sheet.add_chart(chart, "E2")
