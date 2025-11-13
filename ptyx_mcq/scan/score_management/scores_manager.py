import csv
import math
from collections import Counter
from typing import TYPE_CHECKING, Callable

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter  # type: ignore
from openpyxl.chart import Reference, LineChart  # type: ignore
from openpyxl.chart.layout import Layout, ManualLayout  # type: ignore
from openpyxl.drawing.line import LineProperties  # type: ignore

from ptyx.pretty_print import print_info, term_color, TermColors, red, green, yellow, bold

from ptyx_mcq.scan.score_management.evaluation_strategies import (
    AnswersData,
    ScoreData,
    EvaluationStrategies,
)
from ptyx_mcq.tools.config_parser import StudentId, StudentName

if TYPE_CHECKING:
    from ptyx_mcq.scan.scan_doc import MCQPictureParser

STATISTIC_INFO = {"mean": "AVERAGE", "min": "MIN", "max": "MAX"}


class ScoresManager:
    def __init__(self, mcq_parser: "MCQPictureParser"):
        self.mcq_parser = mcq_parser
        self.scores: dict[tuple[StudentId, StudentName], float | str] = {}
        self.results: dict[tuple[StudentId, StudentName], float] = {}

    @property
    def max_score(self):
        return self.mcq_parser.config.max_score

    def calculate_scores(self) -> None:
        cfg = self.mcq_parser.config
        default_mode = cfg.mode["default"]
        default_weight = cfg.weight["default"]
        default_correct = cfg.correct["default"]
        default_incorrect = cfg.incorrect["default"]
        default_skipped = cfg.skipped["default"]
        default_floor = cfg.floor["default"]
        default_ceil = cfg.ceil["default"]

        for doc_id, doc in self.mcq_parser.scan_data.used_docs.items():
            print(f"Test {doc_id} - {doc.student_name}")
            for q, question in doc.questions.items():
                answered = {answer.answer_num for answer in question if answer.checked}
                correct_ones = {answer.answer_num for answer in question if answer.is_correct}
                neutralized_ones = {answer.answer_num for answer in question if answer.neutralized}
                all_answers = {answer.answer_num for answer in question}

                # Neutralized answers must be removed from each set of answers.
                # (Typically, neutralized answers are answers which were detected faulty during or after the
                # examination).
                answered -= neutralized_ones
                correct_ones -= neutralized_ones
                all_answers -= neutralized_ones

                mode = cfg.mode.get(q, default_mode)

                if mode == "skip":
                    # Used mostly to skip bogus questions.
                    print(f"Question {q} skipped...")
                    continue

                try:
                    func = getattr(EvaluationStrategies, mode)
                except AttributeError:
                    raise AttributeError(f"Unknown evaluation mode: {mode!r}.")

                ans_data = AnswersData(checked=answered, correct=correct_ones, all=all_answers)
                scores_data = ScoreData(
                    correct=float(cfg.correct.get(q, default_correct)),
                    skipped=float(cfg.skipped.get(q, default_skipped)),
                    incorrect=float(cfg.incorrect.get(q, default_incorrect)),
                )
                earn = func(ans_data, scores_data)

                floor = cfg.floor.get(q, default_floor)
                assert floor is None or isinstance(floor, (float, int))
                if floor is not None and earn < floor:
                    earn = floor
                ceil = cfg.ceil.get(q, default_ceil)
                assert ceil is None or isinstance(ceil, (float, int))
                if ceil is not None and earn > ceil:
                    earn = ceil

                if earn == scores_data.correct:
                    color = TermColors.GREEN
                elif earn == scores_data.incorrect:
                    color = TermColors.RED
                else:
                    color = TermColors.YELLOW
                print("-  " + term_color(f"Rating (Q{q}): {earn:g}", color=color))
                # Don't forget to include the weight of the question to calculate the global score.
                earn *= float(cfg.weight.get(q, default_weight))
                # Don't use weight for per question score, since it would make success rates
                # harder to compare.
                question.score = earn

        default = self.mcq_parser.config.default_score
        self.scores = {
            (student_id, student_name): default
            for student_id, student_name in self.mcq_parser.config.students_ids.items()
        }
        self.results = {
            (doc.student_id, doc.student_name): doc.score
            for doc in self.mcq_parser.scan_data
            if doc.score is not None
        }
        self.scores.update(self.results)

    @staticmethod
    def _convert(num: float | str, factor=1.0):
        if isinstance(num, float):
            num = round(factor * num, 2)
        return num

    def print_scores(self) -> None:
        min_score: float = math.inf
        max_score: float = -math.inf
        print()
        print(term_color(f"SCORES (/{self.max_score:g}):", color=TermColors.CYAN))
        for (student_id, student_name), score in sorted(self.scores.items()):
            score = self._convert(score)
            if isinstance(score, (float, int)):
                if score < min_score:
                    min_score = score
                if score > max_score:
                    max_score = score
            print(f" - {student_name}: {self._convert(score)}")
        if self.results:
            mean = round(sum(self.results.values()) / len(self.results), 2)
            print(yellow("Mean: " + bold(f"{mean:g}") + f"/{self.max_score:g}"))
            print(f"Min: {red(min_score)} - Max: {green(max_score)}")
        else:
            print("No score found !")
        print()

    def _write_scores(self, writerow: Callable) -> None:
        # max_score is the maximal theoretical score, not the highest actually obtained.
        max_score = self.max_score
        writerow(("ID", "Name", f"Score/{max_score:g}", "Score/20", "Score/100"))
        for (student_id, student_name), score in sorted(self.scores.items()):
            # TODO: Add ability to change the notation system.
            writerow(
                [
                    student_id,
                    student_name,
                    self._convert(score),
                    self._convert(score, factor=20 / max_score) if max_score else 0,
                    self._convert(score, factor=100 / max_score) if max_score else 0,
                ]
            )

    def _append_chart(self, wb: Workbook) -> None:
        """Append a chart representing the scores' distribution as a new sheet."""
        ws = wb.create_sheet(title="Scores Distribution")

        # Count how many times each score appears
        score_counts = Counter(
            (round(score) for score in self.scores.values() if isinstance(score, (int, float)))
        )

        # Prepare continuous range 0–<max-score> (even if some scores are missing)
        ws.append(["Score", "Count"])
        for score in range(math.ceil(self.max_score) + 1):
            ws.append([score, score_counts.get(score, 0)])

        # Data range for chart
        n = ws.max_row
        counts_ref = Reference(ws, min_col=2, min_row=2, max_row=n)  # counts
        scores_ref = Reference(ws, min_col=1, min_row=2, max_row=n)  # scores

        # Create a line chart
        chart = LineChart()
        chart.title = "Scores Distribution"
        chart.x_axis.title = "Score"
        chart.y_axis.title = "Number of Students"
        chart.style = 13

        # Add data and categories
        chart.add_data(counts_ref, titles_from_data=False)
        chart.set_categories(scores_ref)
        # chart.xvalues = scores_ref

        chart.style = 2
        # chart.legend.overlay = False
        chart.layout = Layout()
        chart.x_axis.delete = False
        chart.y_axis.delete = False
        chart.width = 21
        chart.height = 9
        chart.layout = Layout(
            manualLayout=ManualLayout(x=0.005, y=0.05, w=0.75, h=0.8, xMode="factor", yMode="factor")
        )
        chart.layout.layoutTarget = "inner"
        chart.legend = None
        chart.series[0].smooth = False
        # Configure axes
        chart.x_axis.scaling.min = 1
        chart.x_axis.scaling.max = self.max_score + 1

        # Set line color to solid red
        chart.series[0].graphicalProperties.line = LineProperties(
            solidFill="FF0000",  # RGB hex code for red
        )

        # Add the chart to the worksheet
        ws.add_chart(chart, "E2")

    def generate_csv_file(self) -> None:
        scores_path = self.mcq_parser.scan_data.files.csv_scores
        with open(scores_path, "w", newline="") as csvfile:
            self._write_scores(writerow=csv.writer(csvfile).writerow)
        print_info(f'Results stored in "{scores_path}"')

    def generate_xlsx_file(self) -> None:
        wb = Workbook()
        # grab the active worksheet
        sheet: Worksheet = wb.active
        sheet.title = "Resume"
        self._write_scores(writerow=sheet.append)
        tab = Table(displayName="Table1", ref=f"A1:{get_column_letter(sheet.max_column)}{sheet.max_row}")

        # Add a default style with striped rows and banded columns
        style = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=True,
        )
        # noinspection PyTypeChecker
        tab.tableStyleInfo = style
        sheet.column_dimensions["A"].width = 1.23 * max(len(student_id) for student_id, _ in self.scores)
        sheet.column_dimensions["B"].width = 1.23 * max(len(student_name) for _, student_name in self.scores)
        sheet.append([])
        # Append some statistics concerning the students' scores.
        n = len(self.scores)
        for legend, formula in STATISTIC_INFO.items():
            sheet.append([legend.capitalize(), ""] + [f"={formula}({col}2:{col}{n + 1})" for col in "CDE"])

        self._append_chart(wb)
        wb.save(self.mcq_parser.scan_data.files.xlsx_scores)
