import csv
import math
from typing import TYPE_CHECKING, Callable

from ptyx_mcq.scan.evaluation_strategies import AnswersData, ScoreData, EvaluationStrategies
from ptyx_mcq.tools.io_tools import (
    print_info,
    ANSI_RESET,
    ANSI_RED,
    ANSI_GREEN,
    ANSI_YELLOW,
    ANSI_CYAN,
    ANSI_BOLD,
)

if TYPE_CHECKING:
    from ptyx_mcq.scan.scan import MCQPictureParser


class ScoresManager:
    def __init__(self, mcq_parser: "MCQPictureParser"):
        self.mcq_parser = mcq_parser
        self.scores: dict[str, float | str] = {}
        self.results: dict[str, float] = {}

    @property
    def max_score(self):
        return self.mcq_parser.config.max_score

    def calculate_scores(self) -> None:
        cfg = self.mcq_parser.config
        default_mode = cfg.mode["default"]
        default_weight = cfg.weight["default"]
        default_correct = cfg.correct["default"]
        default_incorrect = cfg.incorrect["default"]
        default_skipped = cfg.skipped["default"]
        default_floor = cfg.floor["default"]
        default_ceil = cfg.ceil["default"]

        data_manager = self.mcq_parser.data_storage
        for doc_id in self.mcq_parser.data:
            correct_ans = data_manager.correct_answers[doc_id]
            neutralized_ans = data_manager.neutralized_answers[doc_id]
            doc_data = self.mcq_parser.data[doc_id]
            print(f'Test {doc_id} - {doc_data["name"]}')
            for q in sorted(doc_data["answered"]):
                answered = set(doc_data["answered"][q])
                correct_ones = correct_ans[q]
                neutralized_ones = neutralized_ans[q]
                all_answers = {ans_num for ans_num, is_ok in cfg.ordering[doc_id]["answers"][q]}

                # Neutralized answers must be removed from each set of answers.
                # (Typically, neutralized answers are answers which were detected faulty during or after the
                # examination).
                answered -= neutralized_ones
                correct_ones -= neutralized_ones
                all_answers -= neutralized_ones

                mode = cfg.mode.get(q, default_mode)

                if mode == "skip":
                    # Used mostly to skip bogus questions.
                    print(f"Question {q} skipped...")
                    continue

                try:
                    func = getattr(EvaluationStrategies, mode)
                except AttributeError:
                    raise AttributeError(f"Unknown evaluation mode: {mode!r}.")

                ans_data = AnswersData(checked=answered, correct=correct_ones, all=all_answers)
                scores_data = ScoreData(
                    correct=float(cfg.correct.get(q, default_correct)),
                    skipped=float(cfg.skipped.get(q, default_skipped)),
                    incorrect=float(cfg.incorrect.get(q, default_incorrect)),
                )
                earn = func(ans_data, scores_data)

                floor = cfg.floor.get(q, default_floor)
                assert floor is None or isinstance(floor, (float, int))
                if floor is not None and earn < floor:
                    earn = floor
                ceil = cfg.ceil.get(q, default_ceil)
                assert ceil is None or isinstance(ceil, (float, int))
                if ceil is not None and earn > ceil:
                    earn = ceil

                if earn == scores_data.correct:
                    color = ANSI_GREEN
                elif earn == scores_data.incorrect:
                    color = ANSI_RED
                else:
                    color = ANSI_YELLOW
                print(f"-  {color}Rating (Q{q}): {color}{earn:g}{ANSI_RESET}")
                # Don't forget to include the weight of the question to calculate the global score.
                earn *= float(cfg.weight.get(q, default_weight))
                # Don't use weight for per question score, since it would make success rates
                # harder to compare.
                doc_data["score_per_question"][q] = earn
                doc_data["score"] += earn
        default = self.mcq_parser.config.default_score
        self.scores = {name: default for name in self.mcq_parser.config.students_ids.values()}
        self.results = {doc_data["name"]: doc_data["score"] for doc_data in self.mcq_parser.data.values()}
        self.scores.update(self.results)

    @staticmethod
    def _convert(num: float | str, factor=1.0):
        if isinstance(num, float):
            num = round(factor * num, 2)
        return num

    def print_scores(self) -> None:
        min_score: float = math.inf
        max_score: float = -math.inf
        print()
        print(f"{ANSI_CYAN}SCORES (/{self.max_score:g}):{ANSI_RESET}")
        for name, score in sorted(self.scores.items()):
            score = self._convert(score)
            if isinstance(score, (float, int)):
                if score < min_score:
                    min_score = score
                if score > max_score:
                    max_score = score
            print(f" - {name}: {self._convert(score)}")
        if self.results:
            mean = round(sum(self.results.values()) / len(self.results), 2)
            print(
                f"{ANSI_YELLOW}Mean: "
                f"{ANSI_BOLD}{mean:g}{ANSI_RESET}{ANSI_YELLOW}/{self.max_score:g}{ANSI_RESET}"
            )
            print(f"Min: {ANSI_RED}{min_score}{ANSI_RESET} - Max: {ANSI_GREEN}{max_score}{ANSI_RESET}")
        else:
            print("No score found !")
        print()

    def _write_scores(self, writerow: Callable) -> None:
        max_score = self.max_score
        writerow(("Name", f"Score/{max_score:g}", "Score/20", "Score/100"))
        for name, score in sorted(self.scores.items()):
            # TODO: Add ability to change the notation system.
            writerow(
                [
                    name,
                    self._convert(score),
                    self._convert(score, factor=20 / max_score) if max_score else 0,
                    self._convert(score, factor=100 / max_score) if max_score else 0,
                ]
            )

    def generate_csv_file(self) -> None:
        scores_path = self.mcq_parser.data_storage.files.csv_scores
        with open(scores_path, "w", newline="") as csvfile:
            self._write_scores(writerow=csv.writer(csvfile).writerow)
        print_info(f'Results stored in "{scores_path}"')

    def generate_xlsx_file(self) -> None:
        from openpyxl import Workbook
        from openpyxl.worksheet.worksheet import Worksheet
        from openpyxl.worksheet.table import Table, TableStyleInfo
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        # grab the active worksheet
        sheet: Worksheet = wb.active  # type: ignore
        sheet.title = "Resume"
        self._write_scores(writerow=sheet.append)
        tab = Table(displayName="Table1", ref=f"A1:{get_column_letter(sheet.max_column)}{sheet.max_row}")

        # Add a default style with striped rows and banded columns
        style = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=True,
        )
        tab.tableStyleInfo = style
        sheet.column_dimensions["A"].width = 1.23 * max(len(name) for name in self.scores)
        wb.save(self.mcq_parser.data_storage.files.xlsx_scores)
